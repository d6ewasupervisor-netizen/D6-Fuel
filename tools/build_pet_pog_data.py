"""
Build client-side planogram data from FM_Frozen_Refrig_Pet_Pog_Positions_*.xlsx

Outputs:
  pet_pog_app/data/planograms.json
  pet_pog_app/js/planograms-data.js  (window.PET_POG_PLANOGRAMS; works with file://)

Run from repo root:
  python tools/build_pet_pog_data.py
"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
XLSX = REPO / "FM_Frozen_Refrig_Pet_Pog_Positions_P02W04Y2026.xlsx"
OUT_JSON = REPO / "pet_pog_app" / "data" / "planograms.json"
OUT_JS = REPO / "pet_pog_app" / "js" / "planograms-data.js"

SKIP_SHEETS = frozenset({"Summary", "Master List"})


def parse_width_ft(meta: str) -> float:
    """e.g. '... |   3 ft 6 in W  x  51 in H ...' -> 3.5"""
    m = re.search(r"(\d+(?:\.\d+)?)\s*ft\s*(\d+(?:\.\d+)?)\s*in\s*W", meta or "", re.I)
    if not m:
        return 3.5
    return float(m.group(1)) + float(m.group(2)) / 12.0


def normalize_upc(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        v = int(v)
    if isinstance(v, int):
        s = str(v)
    else:
        s = str(v).strip()
        if s.endswith(".0"):
            s = s[:-2]
    digits = "".join(c for c in s if c.isdigit())
    if not digits:
        return ""
    if len(digits) <= 13:
        return digits.zfill(13)
    return digits


def sheet_label(first_cell: str | None) -> str:
    if not first_cell:
        return "Planogram"
    return str(first_cell).strip()


def parse_planogram_sheet(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    title = sheet_label(rows[0][0] if rows else None)
    meta = str(rows[1][0] or "") if len(rows) > 1 else ""
    dbkey_m = re.search(r"DBKey\s+(\d+)", meta)
    dbkey = dbkey_m.group(1) if dbkey_m else ""
    width_ft = parse_width_ft(meta)

    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == "Bay" and row[1] == "Shelf":
            header_idx = i
            break
    if header_idx is None:
        return {
            "planogram_dbkey": dbkey,
            "label": title,
            "width_ft": width_ft,
            "bays": [],
        }

    # bay -> shelf -> list of products (in file order per shelf)
    bays: dict[int, dict[int, list]] = defaultdict(lambda: defaultdict(list))
    current_bay: int | None = None

    for row in rows[header_idx + 1 :]:
        if not row or row[0] is None:
            continue
        a0 = row[0]
        if isinstance(a0, str) and a0.strip().upper().startswith("BAY"):
            m = re.search(r"(\d+)", a0)
            current_bay = int(m.group(1)) if m else None
            continue
        if current_bay is None:
            continue
        if not isinstance(a0, (int, float)):
            continue

        bay = int(a0)
        shelf = int(row[1])
        position = int(row[2])
        upc = normalize_upc(row[4])
        desc = str(row[5] or "").strip()
        size = str(row[6] or "").strip()
        facings = int(row[7]) if row[7] is not None else 1
        h = float(row[8]) if row[8] is not None else 5.0
        w = float(row[9]) if row[9] is not None else 2.5
        status = str(row[10] or "").strip().upper() if len(row) > 10 and row[10] else ""
        is_new = status == "NEW"
        is_changed = status in {"CHANGED", "CHANGE"}

        bays[bay][shelf].append(
            {
                "upc": upc,
                "description": desc,
                "full_name": desc,
                "bay": bay,
                "shelf": shelf,
                "position": position,
                "size": size,
                "facings": facings,
                "height_inches": h,
                "width_inches": w,
                "is_new": is_new,
                "is_changed": is_changed,
            }
        )

    bay_list = []
    for bay_num in sorted(bays.keys()):
        shelf_map = bays[bay_num]
        shelves_out = []
        for shelf_num in sorted(shelf_map.keys()):
            prods = shelf_map[shelf_num]
            prods.sort(key=lambda p: (p["position"], p["upc"]))
            shelves_out.append({"shelf": shelf_num, "products": prods})
        bay_list.append({"bay": bay_num, "width_ft": width_ft, "shelves": shelves_out})

    return {
        "planogram_dbkey": dbkey,
        "label": title,
        "width_ft": width_ft,
        "bays": bay_list,
    }


def main():
    if not XLSX.exists():
        raise SystemExit(f"Missing workbook: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    planograms = []
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        pog = parse_planogram_sheet(wb[name])
        pog["sheet_name"] = name
        planograms.append(pog)

    payload = {
        "period": "P02 W04 Y2026",
        "effective": "2026-03-22",
        "planograms": planograms,
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JS.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    js_body = json.dumps(payload, indent=2, ensure_ascii=True)
    OUT_JS.write_text(
        "/* Generated by tools/build_pet_pog_data.py - do not edit */\n"
        f"window.PET_POG_PLANOGRAMS = {js_body};\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUT_JSON.relative_to(REPO)}")
    print(f"Wrote {OUT_JS.relative_to(REPO)}")
    print(f"Planograms: {len(planograms)}")


if __name__ == "__main__":
    main()
